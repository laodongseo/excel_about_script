# ‐*‐coding:utf‐8‐*‐
"""
读取某个目录下多个excel
读取每个excel重新生成1下
"""
import os
import pandas
import re
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


# 获取文件集(含路径),ext为文件后缀(带点)
def get_files(file_path, ext):
    file_list = []
    dir_or_files = os.listdir(file_path)
    # dir_or_file不带路径
    for dir_or_file in dir_or_files:
        # 添加路径
        dir_file_path = os.path.join(file_path, dir_or_file)
        # 保留文件,去除文件夹
        if not os.path.isdir(dir_file_path):
            if os.path.splitext(dir_file_path)[-1] == ext:
                file_list.append(dir_file_path)
    return file_list



def make_file(files):
    for file in files:
        wb = load_workbook(file)
        sheet_name = wb.sheetnames[0]
        data = wb[sheet_name].rows
        base_name = os.path.basename(file)
        new_path = f'{res_path}{base_name}'
        wb_last = Workbook()
        ws = wb_last.active
        for row in data:
            new_row = []
            for cell in row:
                new_row.append(cell.value)
            ws.append(new_row)
        print(new_path)
        wb_last.save(new_path)


if __name__ == "__main__":
    excel_path = r"./QM周度/"
    res_path = r'./QM周度_new/'
    files = get_files(excel_path,'.xlsx')
    make_file(files)
